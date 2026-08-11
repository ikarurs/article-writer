#!/usr/bin/env python3
"""Reproduce the public-data first version of the article analysis.

This script intentionally does not estimate an energy-price effect. The required
GLES microdata and licensed historical Tankerkoenig prices were unavailable at
the time of retrieval; see sources/source-log.md.
"""

from __future__ import annotations

import csv
import io
import json
import os
import sqlite3
import tempfile
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from statistics import median, quantiles

import geopandas as gpd
import matplotlib.pyplot as plt
import pandas as pd
import spur
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from shapely.geometry import Point


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
PROCESSED = ROOT / "data" / "processed"
FIGURES = ROOT / "outputs" / "figures"
TABLES = ROOT / "outputs" / "tables"

PARTIES = ("SPD", "CDU", "Greens", "FDP", "AfD", "Left")
COLORS = {
    "SPD": "#E3000F",
    "CDU": "#333333",
    "Greens": "#64A12D",
    "FDP": "#FFED00",
    "AfD": "#009EE0",
    "Left": "#BE3075",
}


def phase(day: date) -> str:
    if day <= date(2022, 2, 23):
        return "Pre-war"
    if day <= date(2022, 6, 2):
        return "Invasion"
    if day <= date(2022, 12, 31):
        return "Anticipation/high-price"
    if day <= date(2023, 12, 31):
        return "Realized supply break"
    return "Later adjustment"


def as_int(value: str | int | float | None, label: str) -> int:
    if value in (None, ""):
        raise ValueError(f"Missing required value: {label}")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Expected integer for {label}, got {value!r}") from exc


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def wsi_series() -> list[dict]:
    # Figure 1, report page 5. October 2021 is recalled actual second vote;
    # the other observations are the published Sunday-question response.
    observations = [
        (date(2021, 7, 1), "July 2021", "Vote intention", 11.0, 4531),
        (date(2021, 10, 1), "October 2021", "Actual Bundestag second vote", 9.0, 4464),
        (date(2022, 1, 1), "January 2022", "Vote intention", 12.0, 4696),
        (date(2022, 11, 1), "November 2022", "Vote intention", 16.0, 3748),
        (date(2023, 7, 1), "July 2023", "Vote intention", 23.0, 3727),
    ]
    rows = []
    for day, label, measure, value, sample in observations:
        if not 0 <= value <= 100:
            raise ValueError("WSI percentage outside 0-100")
        rows.append(
            {
                "date": day.isoformat(),
                "period_label": label,
                "phase": phase(day),
                "measure": measure,
                "party": "AfD",
                "percent": f"{value:.1f}",
                "sample_n": sample,
                "sample_description": "WSI labour-force online quota panel; published weighted percentage",
                "source": "Hoevermann (2023), WSI Report 92, Figure 1, report p. 5",
            }
        )
    write_csv(
        PROCESSED / "wsi_timeline.csv",
        list(rows[0]),
        rows,
    )
    return rows


def read_ni(path: Path, year: int) -> dict[str, dict]:
    party_columns = {
        2017: {
            "SPD": "SPD II",
            "CDU": "CDU II",
            "Greens": "GRÜNE II",
            "FDP": "FDP II",
            "AfD": "AfD Niedersachsen II",
            "Left": "DIE LINKE. II",
        },
        2022: {
            "SPD": "SPD II",
            "CDU": "CDU II",
            "Greens": "GRÜNE II",
            "FDP": "FDP II",
            "AfD": "AfD II",
            "Left": "DIE LINKE. II",
        },
    }[year]
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        required = {"Wahlkreis", "Wahlberechtigte", "Wähler", "Gültige Zweitstimmen", *party_columns.values()}
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"{path.name}: missing columns {sorted(missing)}")
        result: dict[str, dict] = {}
        for row in reader:
            unit = row["Wahlkreis"].strip()
            if unit in result:
                raise ValueError(f"{path.name}: duplicate constituency {unit}")
            eligible = as_int(row["Wahlberechtigte"], f"{unit} eligible")
            voters = as_int(row["Wähler"], f"{unit} voters")
            valid = as_int(row["Gültige Zweitstimmen"], f"{unit} valid votes")
            values = {
                "eligible": eligible,
                "voters": voters,
                "valid": valid,
                **{party: as_int(row[col], f"{unit} {party}") for party, col in party_columns.items()},
            }
            result[unit] = values
    return result


def validate_share(value: float, label: str) -> float:
    if not 0 <= value <= 100:
        raise ValueError(f"{label}: share outside 0-100 ({value})")
    return value


def validate_counts(values: dict, label: str, *, multi_vote: bool = False) -> None:
    if values["eligible"] <= 0 or values["valid"] <= 0:
        raise ValueError(f"{label}: nonpositive denominator")
    if values["voters"] < 0 or values["voters"] > values["eligible"]:
        raise ValueError(f"{label}: voters outside 0..eligible")
    if not multi_vote and values["valid"] > values["voters"]:
        raise ValueError(f"{label}: valid votes exceed voters")
    for party in PARTIES:
        if values[party] < 0 or values[party] > values["valid"]:
            raise ValueError(f"{label}: {party} votes outside 0..valid")


def lower_saxony() -> tuple[list[dict], list[dict]]:
    old = read_ni(RAW / "elections" / "state" / "ni_landtag_2017.csv", 2017)
    new = read_ni(RAW / "elections" / "state" / "ni_landtag_2022.csv", 2022)
    common = sorted(set(old) & set(new))
    if not common:
        raise ValueError("Lower Saxony: no common constituency IDs")

    changes: list[dict] = []
    for unit in common:
        before, after = old[unit], new[unit]
        row = {"constituency_id": unit}
        for party in PARTIES:
            p17 = validate_share(100 * before[party] / before["valid"], f"2017 {unit} {party}")
            p22 = validate_share(100 * after[party] / after["valid"], f"2022 {unit} {party}")
            row[f"{party}_2017_pct"] = f"{p17:.6f}"
            row[f"{party}_2022_pct"] = f"{p22:.6f}"
            row[f"{party}_change_pp"] = f"{p22 - p17:.6f}"
        turnout17 = validate_share(100 * before["voters"] / before["eligible"], f"2017 {unit} turnout")
        turnout22 = validate_share(100 * after["voters"] / after["eligible"], f"2022 {unit} turnout")
        row["turnout_2017_pct"] = f"{turnout17:.6f}"
        row["turnout_2022_pct"] = f"{turnout22:.6f}"
        row["turnout_change_pp"] = f"{turnout22 - turnout17:.6f}"
        changes.append(row)

    write_csv(PROCESSED / "lower_saxony_constituency_changes.csv", list(changes[0]), changes)

    # Independent statewide totals from the official final-result reports
    # (LW/000.pdf), rather than totals re-used from the constituency files.
    official = {
        2017: {"eligible": 6098379, "voters": 3848865, "valid": 3827850,
               "SPD": 1413846, "CDU": 1287191, "Greens": 334131,
               "FDP": 287957, "AfD": 235853, "Left": 177118},
        2022: {"eligible": 6064738, "voters": 3657967, "valid": 3623886,
               "SPD": 1211447, "CDU": 1017304, "Greens": 526940,
               "FDP": 170303, "AfD": 396844, "Left": 98586},
    }
    validation = []
    for year, data in ((2017, old), (2022, new)):
        for measure in ("eligible", "voters", "valid", *PARTIES):
            total = sum(values[measure] for values in data.values())
            difference = total - official[year][measure]
            if difference:
                raise ValueError(f"Lower Saxony {year} {measure}: official-total difference {difference}")
            validation.append(
                {
                    "election": f"Lower Saxony Landtag {year}",
                    "measure": measure,
                    "source_total": official[year][measure],
                    "recomputed_total": total,
                    "difference": difference,
                    "units": len(data),
                    "excluded_units": 0,
                    "validation_basis": "Independent official statewide final-result report LW/000.pdf",
                }
            )
    return changes, validation


def berlin_aggregate() -> tuple[list[dict], list[dict]]:
    path = RAW / "elections" / "state" / "berlin_agh_2021_2023.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "AGH_W2" not in workbook.sheetnames:
        raise ValueError("Berlin workbook: missing AGH_W2 sheet")
    sheet = workbook["AGH_W2"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    required = [
        "Wahlberechtigte insgesamt",
        "Wählende",
        "Gültige Stimmen",
        "SPD",
        "CDU",
        "GRÜNE",
        "DIE LINKE",
        "AfD",
        "FDP",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Berlin workbook: missing columns {missing}")
    index = {name: headers.index(name) for name in required}
    totals = {name: 0 for name in required}
    unit_count = 0
    for row in rows:
        unit_count += 1
        for name, column in index.items():
            totals[name] += as_int(row[column] or 0, f"Berlin row {unit_count} {name}")

    # Official rounded percentages from the AfS Berlin-Brandenburg historical
    # series. The downloadable 2023 workbook contains 2023 raw counts only;
    # archived granular 2021 data could not be recovered reproducibly.
    official_2021 = {"SPD": 21.4, "CDU": 18.0, "Greens": 18.9, "FDP": 7.1, "AfD": 8.0, "Left": 14.1}
    columns = {"SPD": "SPD", "CDU": "CDU", "Greens": "GRÜNE", "FDP": "FDP", "AfD": "AfD", "Left": "DIE LINKE"}
    result = []
    for party in PARTIES:
        share_2023 = validate_share(100 * totals[columns[party]] / totals["Gültige Stimmen"], f"Berlin 2023 {party}")
        result.append(
            {
                "party": party,
                "share_2021_pct_official_rounded": f"{official_2021[party]:.1f}",
                "share_2023_pct_recomputed": f"{share_2023:.6f}",
                "change_pp_approximate": f"{share_2023 - official_2021[party]:.6f}",
                "geography": "Berlin aggregate",
                "note": "Approximate change because the official 2021 input is rounded to 0.1 percentage point",
            }
        )
    write_csv(PROCESSED / "berlin_aggregate_changes.csv", list(result[0]), result)

    official_2023_rounded = {"SPD": 18.4, "CDU": 28.2, "Greens": 18.4, "FDP": 4.6, "AfD": 9.1, "Left": 12.2}
    validation = []
    for party in PARTIES:
        recomputed = 100 * totals[columns[party]] / totals["Gültige Stimmen"]
        difference = recomputed - official_2023_rounded[party]
        if abs(difference) > 0.051:
            raise ValueError(f"Berlin 2023 {party}: aggregate differs from official rounded share by {difference:.3f} pp")
        validation.append(
            {
                "election": "Berlin Abgeordnetenhaus 2023 repeat",
                "measure": f"{party} list-vote share (pct)",
                "source_total": f"{official_2023_rounded[party]:.1f}",
                "recomputed_total": f"{recomputed:.6f}",
                "difference": f"{difference:.6f}",
                "units": unit_count,
                "excluded_units": 0,
                "validation_basis": "All ordinary and postal rows; official percentage rounded to 0.1 pp",
            }
        )
    turnout = 100 * totals["Wählende"] / totals["Wahlberechtigte insgesamt"]
    validation.append(
        {
            "election": "Berlin Abgeordnetenhaus 2023 repeat",
            "measure": "turnout (pct)",
            "source_total": "62.9",
            "recomputed_total": f"{turnout:.6f}",
            "difference": f"{turnout - 62.9:.6f}",
            "units": unit_count,
            "excluded_units": 0,
            "validation_basis": "All ordinary and postal rows; official percentage rounded to 0.1 pp",
        }
    )
    return result, validation


def sh_ags(region_code: str) -> str:
    """Convert Schleswig-Holstein Regionalschluessel to the 8-digit AGS."""
    code = str(region_code).strip()
    if len(code) == 5:  # Kreisfreie Stadt in the 2023 file.
        return code + "000"
    core = code[:12]
    if len(core) != 12:
        raise ValueError(f"Unexpected Schleswig-Holstein region code: {region_code!r}")
    return core[:5] + core[-3:]


def aggregate_rows(rows: list[dict], group_for_row, columns: dict[str, str]) -> dict[str, dict[str, int]]:
    out: dict[str, dict[str, int]] = defaultdict(lambda: {name: 0 for name in columns})
    for row in rows:
        group = group_for_row(row)
        if group is None:
            continue
        for name, column in columns.items():
            raw = row.get(column, "")
            out[group][name] += int(raw) if raw not in (None, "") else 0
    return dict(out)


def schleswig_holstein_landtag() -> tuple[list[dict], dict]:
    state = RAW / "elections" / "state"
    with (state / "sh_ltw_2017_precinct.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows17 = list(csv.DictReader(handle, delimiter=";"))
    with (state / "sh_ltw_2022_precinct.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows22 = list(csv.DictReader(handle, delimiter=";"))

    crosswalk: dict[str, str] = {}
    duplicate_ids: set[str] = set()
    for row in rows22:
        precinct = row["Wahlbezirk Nr."]
        ags = sh_ags(row["Gemeinde/Amt Nr."])
        if precinct in crosswalk and crosswalk[precinct] != ags:
            duplicate_ids.add(precinct)
        crosswalk[precinct] = ags
    if duplicate_ids:
        raise ValueError(f"SH 2022 precinct IDs map to multiple municipalities: {sorted(duplicate_ids)[:5]}")

    matched_ids = {row["Wahlbezirk Nr."] for row in rows17 if row["Wahlbezirk Nr."] in crosswalk}
    columns17 = {
        "eligible": "Wahlberechtigte gesamt (A)",
        "voters": "Waehlende gesamt (B)",
        "valid": "Listenstimmen gueltige (F)",
        "CDU": "F1",
        "SPD": "F2",
        "Greens": "F3",
        "FDP": "F4",
        "Left": "F7",
        "AfD": "F10",
    }
    columns22 = {
        "eligible": "Wahlberechtigte gesamt (A)",
        "voters": "Waehlende gesamt (B)",
        "valid": "Listenstimmen gueltige (F)",
        "CDU": "F1",
        "SPD": "F2",
        "Greens": "F3",
        "FDP": "F4",
        "AfD": "F5",
        "Left": "F6",
    }
    agg17 = aggregate_rows(
        [row for row in rows17 if row["Wahlbezirk Nr."] in matched_ids],
        lambda row: crosswalk[row["Wahlbezirk Nr."]],
        columns17,
    )
    agg22 = aggregate_rows(
        [row for row in rows22 if row["Wahlbezirk Nr."] in matched_ids],
        lambda row: sh_ags(row["Gemeinde/Amt Nr."]),
        columns22,
    )
    common = sorted(set(agg17) & set(agg22))
    results = []
    for ags in common:
        old, new = agg17[ags], agg22[ags]
        if old["eligible"] == 0 or new["eligible"] == 0 or old["valid"] == 0 or new["valid"] == 0:
            continue
        validate_counts(old, f"SH Landtag 2017 {ags}")
        validate_counts(new, f"SH Landtag 2022 {ags}")
        row: dict[str, str] = {"ags": ags, "matched_precincts_only": "true"}
        for party in PARTIES:
            p17 = validate_share(100 * old[party] / old["valid"], f"SH 2017 {ags} {party}")
            p22 = validate_share(100 * new[party] / new["valid"], f"SH 2022 {ags} {party}")
            row[f"{party}_2017_pct"] = f"{p17:.6f}"
            row[f"{party}_2022_pct"] = f"{p22:.6f}"
            row[f"{party}_change_pp"] = f"{p22 - p17:.6f}"
        t17 = 100 * old["voters"] / old["eligible"]
        t22 = 100 * new["voters"] / new["eligible"]
        row["turnout_2017_pct"] = f"{t17:.6f}"
        row["turnout_2022_pct"] = f"{t22:.6f}"
        row["turnout_change_pp"] = f"{t22 - t17:.6f}"
        row["eligible_2017_matched"] = str(old["eligible"])
        row["eligible_2022_matched"] = str(new["eligible"])
        results.append(row)
    write_csv(PROCESSED / "sh_landtag_matched_precinct_municipality_changes.csv", list(results[0]), results)
    total_eligible = sum(int(row["Wahlberechtigte gesamt (A)"] or 0) for row in rows17)
    matched_eligible = sum(int(row["Wahlberechtigte gesamt (A)"] or 0) for row in rows17 if row["Wahlbezirk Nr."] in matched_ids)
    meta = {
        "municipalities": len(results),
        "precinct_rows_2017": len(rows17),
        "matched_precinct_ids": len(matched_ids),
        "precinct_match_rate": len(matched_ids) / len(rows17),
        "eligible_coverage_2017": matched_eligible / total_eligible,
    }
    return results, meta


def schleswig_holstein_local() -> tuple[list[dict], dict]:
    state = RAW / "elections" / "state"
    ws = load_workbook(state / "sh_kommunal_2018_municipality.xlsx", read_only=True, data_only=True)["Tabelle1"]
    parties18 = {"CDU": 13, "SPD": 16, "Greens": 19, "FDP": 22, "AfD": 25, "Left": 28}
    old: dict[str, dict] = {}
    for number, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        if row[3] in (None, ""):
            continue
        try:
            ags = str(int(row[3])).zfill(8)
        except (TypeError, ValueError):
            continue
        if ags in old:
            raise ValueError(f"SH local 2018 duplicate AGS {ags}")
        old[ags] = {
            "name": str(row[4]),
            "eligible": as_int(row[8], f"2018 {ags} eligible"),
            "voters": as_int(row[9], f"2018 {ags} voters"),
            "valid": as_int(row[12], f"2018 {ags} valid"),
            **{party: as_int(row[index] or 0, f"2018 {ags} {party}") for party, index in parties18.items()},
        }

    with (state / "sh_kommunal_2023_precinct.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows23 = list(csv.DictReader(handle, delimiter=";"))
    columns23 = {
        "eligible": "Wahlberechtigte gesamt (A)",
        "voters": "Waehlende gesamt (B)",
        "valid": "Stimmen gueltige (D)",
        "CDU": "D1",
        "Greens": "D2",
        "SPD": "D3",
        "FDP": "D4",
        "AfD": "D6",
        "Left": "D7",
    }
    new = aggregate_rows(rows23, lambda row: sh_ags(row["Gemeinde"]), columns23)
    common = sorted(set(old) & set(new))
    results = []
    coverage = {party: 0 for party in PARTIES}
    for ags in common:
        before, after = old[ags], new[ags]
        validate_counts(before, f"SH local 2018 {ags}", multi_vote=True)
        validate_counts(after, f"SH local 2023 {ags}", multi_vote=True)
        row: dict[str, str] = {"ags": ags, "municipality": before["name"]}
        for party in PARTIES:
            ran_both = before[party] > 0 and after[party] > 0
            row[f"{party}_ran_both"] = str(ran_both).lower()
            if ran_both:
                coverage[party] += 1
                p18 = validate_share(100 * before[party] / before["valid"], f"SH local 2018 {ags} {party}")
                p23 = validate_share(100 * after[party] / after["valid"], f"SH local 2023 {ags} {party}")
                row[f"{party}_2018_pct"] = f"{p18:.6f}"
                row[f"{party}_2023_pct"] = f"{p23:.6f}"
                row[f"{party}_change_pp"] = f"{p23 - p18:.6f}"
            else:
                row[f"{party}_2018_pct"] = ""
                row[f"{party}_2023_pct"] = ""
                row[f"{party}_change_pp"] = ""
        t18 = 100 * before["voters"] / before["eligible"]
        t23 = 100 * after["voters"] / after["eligible"]
        row["turnout_2018_pct"] = f"{t18:.6f}"
        row["turnout_2023_pct"] = f"{t23:.6f}"
        row["turnout_change_pp"] = f"{t23 - t18:.6f}"
        results.append(row)
    write_csv(PROCESSED / "sh_local_municipality_changes.csv", list(results[0]), results)
    return results, {"matched_municipalities": len(results), "party_coverage": coverage, "raw_2023_municipalities": len(new)}


def read_brandenburg(path: Path, year: int) -> dict[str, dict]:
    sheet = load_workbook(path, read_only=True, data_only=True)["Brandenburg_Landtagswahl_A_2"]
    rows = sheet.iter_rows(min_row=3, values_only=True)
    if year == 2019:
        indices = {"eligible": 4, "voters": 8, "valid": 11, "SPD": 12, "CDU": 14, "Left": 16, "AfD": 18, "Greens": 20, "FDP": 26}
        types = {"amtsfreie Gemeinde", "amtsangehörige Gemeinde"}
    else:
        indices = {"eligible": 5, "voters": 12, "valid": 16, "SPD": 20, "AfD": 22, "CDU": 24, "Greens": 26, "Left": 28, "FDP": 32}
        types = {"amtsfreie Gemeinde", "amtsangehörige Gemeinde", "verbandsangehörige Gemeinde"}
    result = {}
    for row in rows:
        if row[2] not in types:
            continue
        raw_key = str(row[1])
        if year == 2019:
            ags = raw_key.split()[0]
        else:
            ars = raw_key.removeprefix("SI")
            ags = ars[:5] + ars[-3:]
        if ags in result:
            raise ValueError(f"Brandenburg {year}: duplicate municipality {ags}")
        result[ags] = {"name": str(row[3]), **{name: as_int(row[index], f"BB {year} {ags} {name}") for name, index in indices.items()}}
    return result


def pck_coordinates() -> tuple[float, float]:
    archive = RAW / "spatial" / "prtr_2024.zip"
    with tempfile.TemporaryDirectory() as directory:
        with zipfile.ZipFile(archive) as bundle:
            bundle.extract("prtr_2024.db", directory)
        database = Path(directory) / "prtr_2024.db"
        connection = sqlite3.connect(database)
        try:
            row = connection.execute(
                """SELECT b.ETRS89_x, b.ETRS89_y FROM betriebe b
                   JOIN taetigkeiten t ON t.betriebe_id=b.id
                   WHERE b.jahr=2024 AND b.name LIKE 'PCK Raffinerie%'
                     AND t.prtr_beschreibung='Mineralöl- und Gasraffinerien'
                   LIMIT 1"""
            ).fetchone()
        finally:
            connection.close()
    if not row:
        raise ValueError("PCK Schwedt refinery not found in PRTR 2024")
    return float(row[0]), float(row[1])


def brandenburg_spatial() -> tuple[list[dict], gpd.GeoDataFrame, dict]:
    state = RAW / "elections" / "state"
    old = read_brandenburg(state / "bb_ltw_2019_aggregates.xlsx", 2019)
    new = read_brandenburg(state / "bb_ltw_2024_aggregates.xlsx", 2024)
    common = sorted(set(old) & set(new))
    if len(common) < 400:
        raise ValueError(f"Brandenburg municipality match unexpectedly small: {len(common)}")

    geo = gpd.read_file(RAW / "spatial" / "brandenburg_bbox_vg250.geojson")
    geo = geo.loc[geo["sn_l"].astype(str).eq("12")].copy()
    geo["ags"] = geo["ags"].astype(str).str.zfill(8)
    geo = geo[["ags", "gen", "geometry"]].dissolve(by="ags", aggfunc={"gen": "first"}).reset_index()
    geo_metric = geo.to_crs(25833)
    pck_lon, pck_lat = pck_coordinates()
    pck_metric = gpd.GeoSeries([Point(pck_lon, pck_lat)], crs=4326).to_crs(25833).iloc[0]
    points = geo_metric.geometry.representative_point()
    centroids = geo_metric.geometry.centroid
    geo_metric["lon"] = gpd.GeoSeries(points, crs=25833).to_crs(4326).x
    geo_metric["lat"] = gpd.GeoSeries(points, crs=25833).to_crs(4326).y
    geo_metric["distance_pck_km"] = points.distance(pck_metric) / 1000
    geo_metric["centroid_distance_pck_km"] = centroids.distance(pck_metric) / 1000

    attributes = []
    for ags in common:
        before, after = old[ags], new[ags]
        validate_counts(before, f"Brandenburg Landtag 2019 {ags}")
        validate_counts(after, f"Brandenburg Landtag 2024 {ags}")
        row: dict[str, str | float] = {"ags": ags, "municipality": after["name"]}
        for party in PARTIES:
            p19 = validate_share(100 * before[party] / before["valid"], f"BB 2019 {ags} {party}")
            p24 = validate_share(100 * after[party] / after["valid"], f"BB 2024 {ags} {party}")
            row[f"{party}_2019_pct"] = p19
            row[f"{party}_2024_pct"] = p24
            row[f"{party}_change_pp"] = p24 - p19
        row["turnout_2019_pct"] = 100 * before["voters"] / before["eligible"]
        row["turnout_2024_pct"] = 100 * after["voters"] / after["eligible"]
        row["turnout_change_pp"] = row["turnout_2024_pct"] - row["turnout_2019_pct"]
        row["eligible_2019"] = before["eligible"]
        row["eligible_2024"] = after["eligible"]
        row["valid_2019"] = before["valid"]
        row["valid_2024"] = after["valid"]
        attributes.append(row)
    frame = pd.DataFrame(attributes)
    spatial = geo_metric.merge(frame, on="ags", how="inner", validate="one_to_one")
    if len(spatial) != len(common):
        raise ValueError(f"Brandenburg geography matched {len(spatial)} of {len(common)} election municipalities")
    spatial["distance_pck_km"] = spatial["distance_pck_km"].astype(float)
    output = spatial.drop(columns="geometry").copy()
    output.to_csv(PROCESSED / "brandenburg_municipality_changes_exposure.csv", index=False, encoding="utf-8")

    spatial["distance_quartile"] = pd.qcut(spatial["distance_pck_km"], 4, labels=["Q1 closest", "Q2", "Q3", "Q4 farthest"])
    gradient = (
        spatial.groupby("distance_quartile", observed=True)
        .agg(municipalities=("ags", "count"), mean_distance_km=("distance_pck_km", "mean"), mean_afd_change_pp=("AfD_change_pp", "mean"), median_afd_change_pp=("AfD_change_pp", "median"))
        .reset_index()
    )
    weighted = {}
    for label, group in spatial.groupby("distance_quartile", observed=True):
        weighted[str(label)] = (group["AfD_change_pp"] * group["valid_2024"]).sum() / group["valid_2024"].sum()
    gradient["valid_vote_weighted_mean_afd_change_pp"] = gradient["distance_quartile"].astype(str).map(weighted)
    gradient.to_csv(TABLES / "brandenburg_pck_distance_gradient.csv", index=False, encoding="utf-8")
    sensitivity = pd.DataFrame(
        [
            {"distance_measure": "polygon representative point", "correlation_with_afd_change": spatial["distance_pck_km"].corr(spatial["AfD_change_pp"])},
            {"distance_measure": "polygon centroid", "correlation_with_afd_change": spatial["centroid_distance_pck_km"].corr(spatial["AfD_change_pp"])},
        ]
    )
    sensitivity.to_csv(TABLES / "brandenburg_distance_sensitivity.csv", index=False, encoding="utf-8")
    meta = {"matched_municipalities": len(spatial), "pck_lon": pck_lon, "pck_lat": pck_lat, "geography_year": 2025}
    return output.to_dict("records"), spatial, meta


def brandenburg_spur(spatial: gpd.GeoDataFrame) -> dict:
    """Run the Becker-Boll-Voth SPUR workflow as an exploratory robustness check.

    Proximity is defined as negative representative-point distance in 100 km,
    so a positive levels coefficient would mean a larger AfD change closer to
    PCK. This guards against spatial-unit-root inference; it does not make
    refinery proximity a causal or price exposure.
    """
    frame = spatial[["AfD_change_pp", "distance_pck_km", "lon", "lat"]].copy()
    frame = frame.rename(columns={"AfD_change_pp": "afd_change_pp"})
    frame["proximity_100km"] = -frame["distance_pck_km"] / 100
    result = spur.spur(
        "afd_change_pp ~ proximity_100km",
        frame,
        lon="lon",
        lat="lat",
        q=10,
        nrep=20000,
        seed=42,
    )
    diagnostics = []
    for name in ("i0", "i1", "i0resid", "i1resid"):
        test = getattr(result.tests, name)
        diagnostics.append({"test": name, "lr": test.LR, "p_value": test.pvalue, "q": 10, "nrep": 20000, "seed": 42})
    write_csv(TABLES / "brandenburg_spur_diagnostics.csv", list(diagnostics[0]), diagnostics)

    # Becker, Boll and Voth's 10% rule: use levels only if I(0) is not
    # rejected and I(1) is rejected. Otherwise transform y and x together.
    selected = "levels" if result.tests.i0.pvalue >= 0.10 and result.tests.i1.pvalue < 0.10 else "transformed"
    estimates = []
    for branch_name in ("levels", "transformed"):
        branch = getattr(result.fits, branch_name)
        coefficient_name = "proximity_100km" if branch_name == "levels" else "h_proximity_100km"
        names = branch.scpc.coef_names
        position = names.index(coefficient_name)
        stats = branch.scpc.scpcstats[position]
        estimates.append(
            {
                "branch": branch_name,
                "selected_by_10pct_rule": str(branch_name == selected).lower(),
                "coefficient": float(stats[0]),
                "scpc_se": float(stats[1]),
                "ci95_low": float(stats[4]),
                "ci95_high": float(stats[5]),
                "n": len(frame),
                "scpc_q": branch.scpc.q,
                "scpc_avc": branch.scpc.avc,
                "interpretation_limit": "Exploratory spatial association; representative-point refinery distance is not a measured price shock",
            }
        )
    write_csv(TABLES / "brandenburg_spur_regression.csv", list(estimates[0]), estimates)
    (TABLES / "brandenburg_spur_summary.txt").write_text(result.summary() + "\n", encoding="utf-8")
    return {"selected_branch": selected, "diagnostics": diagnostics, "estimates": estimates}


def load_sh_geography() -> gpd.GeoDataFrame:
    path = RAW / "elections" / "state" / "sh_municipalities_2022.zip"
    with zipfile.ZipFile(path) as outer:
        nested = outer.read("Gemeinden/Gemeinden_WK_LTW_SH_2022_geojson.zip")
    with zipfile.ZipFile(io.BytesIO(nested)) as inner:
        data = json.loads(inner.read("Gemeinden_WK_LTW_SH_2022.geojson"))
    geo = gpd.GeoDataFrame.from_features(data["features"], crs=25832)
    geo["ags"] = geo["ARS"].astype(str).str[:5] + geo["ARS"].astype(str).str[-3:]
    return geo[["ags", "GEM_elect", "geometry"]].dissolve(by="ags", aggfunc={"GEM_elect": "first"}).reset_index()


def plot_spatial(sh_ltw: list[dict], sh_local: list[dict], bb: gpd.GeoDataFrame, pck: tuple[float, float]) -> None:
    FIGURES.mkdir(parents=True, exist_ok=True)
    sh_geo = load_sh_geography()
    ltw = sh_geo.merge(pd.DataFrame(sh_ltw), on="ags", how="left")
    local = sh_geo.merge(pd.DataFrame(sh_local), on="ags", how="left")
    ltw["AfD_change_pp"] = pd.to_numeric(ltw["AfD_change_pp"], errors="coerce")
    local["turnout_change_pp"] = pd.to_numeric(local["turnout_change_pp"], errors="coerce")
    fig, axes = plt.subplots(1, 2, figsize=(13, 7))
    ltw.plot(column="AfD_change_pp", cmap="RdBu_r", legend=True, missing_kwds={"color": "#eeeeee"}, ax=axes[0])
    axes[0].set_title("SH Landtag: AfD change, 2017–May 2022\nmatched stable precincts aggregated to municipalities")
    local.plot(column="turnout_change_pp", cmap="PuOr", legend=True, missing_kwds={"color": "#eeeeee"}, ax=axes[1])
    axes[1].set_title("SH local elections: turnout change, 2018–May 2023\nmatched municipalities")
    for ax in axes:
        ax.axis("off")
    fig.text(0.01, 0.01, "Official Statistik Nord results. Grey denotes unmatched or missing values; maps are descriptive calendar benchmarks.", fontsize=9)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    fig.savefig(FIGURES / "sh_spatial_timing.png", dpi=180)
    plt.close(fig)

    bb_wgs = bb.to_crs(4326)
    fig, axes = plt.subplots(1, 3, figsize=(18, 6.5))
    bb_wgs.plot(column="AfD_change_pp", cmap="RdBu_r", legend=True, ax=axes[0])
    axes[0].set_title("AfD second-vote change\nBrandenburg 2019–2024")
    bb_wgs.plot(column="distance_pck_km", cmap="viridis_r", legend=True, ax=axes[1])
    axes[1].scatter([pck[0]], [pck[1]], marker="*", s=130, color="black", label="PCK Schwedt")
    axes[1].legend(loc="lower left")
    axes[1].set_title("Municipality distance to PCK Schwedt\n(PRTR facility coordinate)")
    axes[2].scatter(bb["distance_pck_km"], bb["AfD_change_pp"], s=12, alpha=0.35, color="#444444")
    bins = bb.groupby(pd.qcut(bb["distance_pck_km"], 10, duplicates="drop"), observed=True).agg(x=("distance_pck_km", "mean"), y=("AfD_change_pp", "mean"))
    axes[2].plot(bins["x"], bins["y"], marker="o", color="#009EE0", linewidth=2)
    axes[2].set_xlabel("Distance to PCK Schwedt (km)")
    axes[2].set_ylabel("AfD change (percentage points)")
    axes[2].set_title("Descriptive distance gradient\npoints and decile means")
    axes[2].grid(alpha=0.2)
    for ax in axes[:2]:
        ax.axis("off")
    fig.text(0.01, 0.01, "Official Brandenburg election returns, BKG municipality geometry, UBA PRTR refinery point. Distance is not a measured price shock.", fontsize=9)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    fig.savefig(FIGURES / "brandenburg_spatial.png", dpi=180)
    plt.close(fig)


def fonts() -> tuple[ImageFont.FreeTypeFont | ImageFont.ImageFont, ...]:
    path = Path("C:/Windows/Fonts/arial.ttf")
    bold = Path("C:/Windows/Fonts/arialbd.ttf")
    if path.exists() and bold.exists():
        return ImageFont.truetype(str(path), 20), ImageFont.truetype(str(path), 16), ImageFont.truetype(str(bold), 28)
    fallback = ImageFont.load_default()
    return fallback, fallback, fallback


def plot_sentiment(rows: list[dict]) -> None:
    width, height = 1500, 850
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    body, small, title = fonts()
    left, right, top, bottom = 115, 80, 105, 155
    plot_w, plot_h = width - left - right, height - top - bottom
    start, end = date(2021, 6, 1), date(2023, 8, 15)
    y_min, y_max = 5.0, 25.0

    def xy(day: date, value: float) -> tuple[float, float]:
        x = left + (day - start).days / (end - start).days * plot_w
        y = top + (y_max - value) / (y_max - y_min) * plot_h
        return x, y

    draw.text((left, 35), "Published WSI AfD support rose before and after January 2023", font=title, fill="#111111")
    for value in range(5, 26, 5):
        _, y = xy(start, value)
        draw.line((left, y, width - right, y), fill="#DDDDDD", width=1)
        draw.text((58, y - 10), f"{value}%", font=small, fill="#444444")
    events = [
        (date(2022, 2, 24), "Invasion"),
        (date(2022, 6, 3), "Oil embargo decided"),
        (date(2023, 1, 1), "Pipeline purchases stop"),
    ]
    for day, label in events:
        x, _ = xy(day, y_min)
        for y in range(top, top + plot_h, 15):
            draw.line((x, y, x, min(y + 7, top + plot_h)), fill="#777777", width=2)
        draw.text((x + 5, top + 10), label, font=small, fill="#555555")
    points = [xy(date.fromisoformat(row["date"]), float(row["percent"])) for row in rows]
    draw.line(points, fill=COLORS["AfD"], width=5)
    for (x, y), row in zip(points, rows):
        draw.ellipse((x - 8, y - 8, x + 8, y + 8), fill=COLORS["AfD"], outline="white", width=2)
        kind = "actual vote" if row["measure"].startswith("Actual") else "intention"
        label = f"{row['period_label']}\n{row['percent']}% ({kind})"
        draw.multiline_text((x - 68, y - 60), label, font=small, fill="#111111", align="center", spacing=2)
    draw.line((left, top + plot_h, width - right, top + plot_h), fill="#333333", width=2)
    draw.line((left, top, left, top + plot_h), fill="#333333", width=2)
    draw.text((left, height - 105), "Source: Hövermann (2023), WSI Report 92, Fig. 1. Labour-force online quota panel; October 2021 is actual vote.", font=small, fill="#333333")
    FIGURES.mkdir(parents=True, exist_ok=True)
    image.save(FIGURES / "sentiment_timeline.png")


def plot_elections(ni_rows: list[dict], berlin_rows: list[dict]) -> None:
    width, height = 1800, 900
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    body, small, title = fonts()
    draw.text((70, 30), "Party vote-share changes in two different state elections", font=title, fill="#111111")
    all_changes = [float(row[f"{party}_change_pp"]) for row in ni_rows for party in PARTIES]
    all_changes.extend(float(row["change_pp_approximate"]) for row in berlin_rows)
    y_min = float(5 * (int(min(all_changes) / 5) - 1))
    y_max = float(5 * (int(max(all_changes) / 5) + 2))
    top, bottom = 145, 170
    plot_h = height - top - bottom

    def ycoord(value: float) -> float:
        return top + (y_max - value) / (y_max - y_min) * plot_h

    panels = [(70, 835), (970, 1730)]
    for panel_left, panel_right in panels:
        for value in range(int(y_min), int(y_max) + 1, 5):
            y = ycoord(value)
            draw.line((panel_left, y, panel_right, y), fill="#E0E0E0", width=1)
            draw.text((panel_left - 48, y - 9), str(value), font=small, fill="#444444")
        draw.line((panel_left, ycoord(0), panel_right, ycoord(0)), fill="#777777", width=2)

    draw.text((120, 90), f"Lower Saxony constituencies, 2017–2022 (n={len(ni_rows)})", font=body, fill="#111111")
    panel_left, panel_right = panels[0]
    step = (panel_right - panel_left) / len(PARTIES)
    for i, party in enumerate(PARTIES):
        values = sorted(float(row[f"{party}_change_pp"]) for row in ni_rows)
        q1, q2, q3 = quantiles(values, n=4, method="inclusive")
        low, high = values[0], values[-1]
        x = panel_left + (i + 0.5) * step
        box_w = 58
        draw.line((x, ycoord(low), x, ycoord(high)), fill="#555555", width=3)
        draw.line((x - 18, ycoord(low), x + 18, ycoord(low)), fill="#555555", width=3)
        draw.line((x - 18, ycoord(high), x + 18, ycoord(high)), fill="#555555", width=3)
        draw.rectangle((x - box_w / 2, ycoord(q3), x + box_w / 2, ycoord(q1)), fill=COLORS[party], outline="#333333", width=2)
        draw.line((x - box_w / 2, ycoord(q2), x + box_w / 2, ycoord(q2)), fill="white", width=4)
        draw.text((x - 30, top + plot_h + 22), party, font=small, fill="#111111")

    draw.text((1000, 90), "Berlin aggregate repeat election, 2021–2023", font=body, fill="#111111")
    panel_left, panel_right = panels[1]
    step = (panel_right - panel_left) / len(PARTIES)
    for i, (party, row) in enumerate(zip(PARTIES, berlin_rows)):
        value = float(row["change_pp_approximate"])
        x = panel_left + (i + 0.5) * step
        zero, end = ycoord(0), ycoord(value)
        draw.rectangle((x - 38, min(zero, end), x + 38, max(zero, end)), fill=COLORS[party], outline="#333333")
        draw.text((x - 30, top + plot_h + 22), party, font=small, fill="#111111")
        draw.text((x - 26, end - 28 if value >= 0 else end + 7), f"{value:+.1f}", font=small, fill="#111111")
    draw.text((70, height - 100), "Percentage-point changes. Sources: official Lower Saxony files and AfS Berlin-Brandenburg. The panels are descriptive and not comparable treatment estimates.", font=small, fill="#333333")
    FIGURES.mkdir(parents=True, exist_ok=True)
    image.save(FIGURES / "election_changes.png")


def timing_table(wsi: list[dict], ni: list[dict], berlin: list[dict]) -> None:
    afd_ni = [float(row["AfD_change_pp"]) for row in ni]
    afd_berlin = next(float(row["change_pp_approximate"]) for row in berlin if row["party"] == "AfD")
    rows = [
        {
            "phase": row["phase"],
            "observation": row["period_label"],
            "geography": "Germany (WSI labour-force panel)",
            "outcome": row["measure"] + ": AfD",
            "value_or_change": row["percent"] + "%",
            "interpretation_limit": "Published aggregate; no local price exposure and no respondent transitions",
        }
        for row in wsi
    ]
    rows.extend(
        [
            {
                "phase": "Anticipation/high-price",
                "observation": "Lower Saxony Landtag, 9 October 2022",
                "geography": f"87 constituencies",
                "outcome": "AfD second-vote share change, 2017–2022",
                "value_or_change": f"median {median(afd_ni):.2f} pp; range {min(afd_ni):.2f} to {max(afd_ni):.2f}",
                "interpretation_limit": "Five-year election change; many concurrent political forces",
            },
            {
                "phase": "Realized supply break",
                "observation": "Berlin repeat election, 12 February 2023",
                "geography": "Berlin aggregate",
                "outcome": "AfD list-vote share change, 2021–2023",
                "value_or_change": f"approximately {afd_berlin:.2f} pp",
                "interpretation_limit": "Court-ordered repeat after administrative failures; rounded 2021 baseline",
            },
        ]
    )
    write_csv(TABLES / "timing_summary.csv", list(rows[0]), rows)


def main() -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    TABLES.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    wsi = wsi_series()
    ni, validation_ni = lower_saxony()
    berlin, validation_berlin = berlin_aggregate()
    sh_ltw, sh_ltw_meta = schleswig_holstein_landtag()
    sh_local, sh_local_meta = schleswig_holstein_local()
    bb_rows, bb_spatial, bb_meta = brandenburg_spatial()
    spur_result = brandenburg_spur(bb_spatial)
    validation = validation_ni + validation_berlin
    write_csv(TABLES / "election_validation.csv", list(validation[0]), validation)
    spatial_coverage = [
        {
            "analysis": "Schleswig-Holstein Landtag 2017-2022",
            "unit": "stable precinct IDs aggregated to municipality",
            "matched_units": sh_ltw_meta["municipalities"],
            "coverage": f"{100 * sh_ltw_meta['eligible_coverage_2017']:.2f}% of 2017 eligible voters",
            "main_limit": "2017 rows without a stable 2022 precinct ID are excluded",
        },
        {
            "analysis": "Schleswig-Holstein local 2018-2023",
            "unit": "municipality",
            "matched_units": sh_local_meta["matched_municipalities"],
            "coverage": "; ".join(f"{p}: {n} ran both" for p, n in sh_local_meta["party_coverage"].items()),
            "main_limit": "party comparisons require candidature in both elections",
        },
        {
            "analysis": "Brandenburg Landtag 2019-2024",
            "unit": "municipality",
            "matched_units": bb_meta["matched_municipalities"],
            "coverage": "409 municipalities common to both elections and matched to BKG geometry",
            "main_limit": "four municipality keys changed between elections; distance to PCK is not a local price shock",
        },
    ]
    write_csv(TABLES / "spatial_coverage.csv", list(spatial_coverage[0]), spatial_coverage)
    timing_table(wsi, ni, berlin)
    plot_sentiment(wsi)
    plot_elections(ni, berlin)
    plot_spatial(sh_ltw, sh_local, bb_spatial, (bb_meta["pck_lon"], bb_meta["pck_lat"]))
    print(
        f"Created {len(wsi)} WSI observations; context results for Lower Saxony and Berlin; "
        f"and spatial panels for {len(sh_ltw)} SH Landtag municipal aggregates, "
        f"{len(sh_local)} SH local-election municipalities, and {len(bb_rows)} Brandenburg municipalities; "
        f"SPUR selected the {spur_result['selected_branch']} branch."
    )


if __name__ == "__main__":
    main()
